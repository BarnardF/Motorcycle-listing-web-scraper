"""
Excel Report Generator
Exports motorcycle listings to Excel with bike model sections
Created with Claude - Jan 2026
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from logger.logger import logger
import json
import os


def generate_excel_report(all_listings, bikes_tracked, output_file="data/listings.xlsx"):
    """
    Generate an Excel report of motorcycle listings with bike model sections
    
    Args:
        all_listings: List of all listings found
        bikes_tracked: List of bike models being tracked
        output_file: Path to output Excel file
    """
    try:
        # Group listings by search term (bike model)
        listings_by_bike = {}
        for listing in all_listings:
            bike = listing.get('search_term', 'Unknown')
            if bike not in listings_by_bike:
                listings_by_bike[bike] = []
            listings_by_bike[bike].append(listing)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "All Listings"
        
        # Setup styles
        header_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        bike_header_fill = PatternFill(start_color="e44c65", end_color="e44c65", fill_type="solid")
        bike_header_font = Font(bold=True, color="FFFFFF", size=10)
        price_drop_fill = PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Column headers
        headers = [
            "Bike Model",
            "Source",
            "Title",
            "Price",
            "Old Price",
            "Kilometers",
            "Location",
            "URL",
            "Search Term",
            "Found Date",
            "Listing ID"
        ]
        
        # Write main header
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Set column widths
        column_widths = [20, 12, 35, 15, 15, 15, 18, 40, 25, 18, 25]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        current_row = 2
        
        # Add listings organized by bike model
        for bike_model in bikes_tracked:
            if bike_model not in listings_by_bike:
                continue
            
            listings = listings_by_bike[bike_model]
            
            # Add bike model label row
            ws.merge_cells(f'A{current_row}:K{current_row}')
            bike_label = ws.cell(row=current_row, column=1)
            bike_label.value = f"{bike_model} ({len(listings)} listings)"
            bike_label.fill = bike_header_fill
            bike_label.font = bike_header_font
            bike_label.alignment = Alignment(horizontal="left", vertical="center")
            bike_label.border = border
            current_row += 1
            
            # Add listings for this bike
            for listing in listings:
                row_data = [
                    bike_model,
                    listing.get('source', ''),
                    listing.get('title', ''),
                    listing.get('price', ''),
                    listing.get('old_price', ''),
                    listing.get('kilometers', ''),
                    listing.get('location', ''),
                    listing.get('url', ''),
                    listing.get('search_term', ''),
                    listing.get('found_date', ''),
                    listing.get('id', '')
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    
                    # Highlight price drops in green
                    if listing.get('price_dropped'):
                        cell.fill = price_drop_fill
                    
                    # Center align certain columns
                    if col_num in [2, 5, 6]:  # Source, Old Price, Kilometers
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                current_row += 1
        
        # Add summary section at bottom
        current_row += 1
        summary_fill = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")
        summary_font = Font(bold=True, size=10)
        
        # Summary label
        ws.merge_cells(f'A{current_row}:B{current_row}')
        summary_cell = ws.cell(row=current_row, column=1)
        summary_cell.value = "SUMMARY"
        summary_cell.fill = summary_fill
        summary_cell.font = summary_font
        summary_cell.border = border
        current_row += 1
        
        # Summary stats
        total_listings = len(all_listings)
        price_drops = sum(1 for l in all_listings if l.get('price_dropped'))
        sources = set(l.get('source') for l in all_listings)
        
        summary_data = [
            ("Total Listings:", total_listings),
            ("Bikes Tracked:", len(bikes_tracked)),
            ("Sources:", len(sources)),
            ("Price Drops Detected:", price_drops),
            ("Last Updated:", datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        ]
        
        for label, value in summary_data:
            ws.cell(row=current_row, column=1).value = label
            ws.cell(row=current_row, column=1).fill = summary_fill
            ws.cell(row=current_row, column=1).font = summary_font
            ws.cell(row=current_row, column=1).border = border
            
            ws.cell(row=current_row, column=2).value = value
            ws.cell(row=current_row, column=2).fill = summary_fill
            ws.cell(row=current_row, column=2).border = border
            
            current_row += 1
        
        # Set row height for header
        ws.row_dimensions[1].height = 25
        
        # Save workbook
        os.makedirs("data", exist_ok=True)
        wb.save(output_file)
        
        logger.info(f"- Generated Excel report: {output_file}")
        logger.info(f" - Total listings: {total_listings}")
        logger.info(f" - Bikes with listings: {len(listings_by_bike)}")
        logger.info(f" - Sources: {len(sources)}")
        if price_drops > 0:
            logger.info(f" - Price drops detected: {price_drops}")
        
        return True
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {e}")
        return False